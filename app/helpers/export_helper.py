import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_absensi(absensi_records, title="REKAPITULASI ABSENSI SISWA SMKN 1 BANGKINANG"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Absensi"

    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    font_header_title = Font(name="Calibri", size=14, bold=True, color="1F2937")
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="4B5563")
    font_tbl_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10, color="111827")
    
    fill_header = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title Headers
    ws.merge_cells("A1:I1")
    ws["A1"] = title
    ws["A1"].font = font_header_title
    ws["A1"].alignment = align_center

    ws.merge_cells("A2:I2")
    ws["A2"] = "SMK NEGERI 1 BANGKINANG - SISTEM ABSENSI DIGITAL GEOFENCING"
    ws["A2"].font = font_subtitle
    ws["A2"].alignment = align_center

    # Empty space
    ws.append([])

    # Table Header Row (Row 4)
    headers = ["No", "Tanggal", "Jam Masuk", "NIS", "Nama Siswa", "Kelas", "Mata Pelajaran", "Status", "Jarak (m)"]
    ws.append(headers)

    header_row_idx = 4
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = font_tbl_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    # Data Rows
    row_start = 5
    for idx, r in enumerate(absensi_records, start=1):
        rec_data = r.to_dict()
        row_values = [
            idx,
            rec_data['tanggal'],
            rec_data['jam_masuk'],
            rec_data['nis'],
            rec_data['nama'],
            rec_data['kelas'],
            rec_data['mata_pelajaran'],
            rec_data['status'],
            f"{rec_data['jarak']} m"
        ]
        ws.append(row_values)
        current_row = row_start + idx - 1

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = font_data
            cell.border = thin_border
            if idx % 2 == 0:
                cell.fill = fill_zebra
            
            if col_idx in [1, 2, 3, 4, 6, 8, 9]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save to memory stream
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
